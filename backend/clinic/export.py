"""Builds a single .xlsx workbook containing EVERYTHING the clinic has ever
recorded for one patient — profile, assessment, follow-up file, progress
log, Mounjaro/Ozempic dose logs, lab tests, prescriptions, nutrition plans
(with meal-level detail), doctor notes, health-status notes, and billing
(invoices + line items). Doctor-only feature (see clinic/views.py::
PatientExportView) — a full data dump including "الحسابات" (billing) isn't
something the secretary should be able to pull.

Pure read/format layer: no model changes, one sheet per section, generated
fresh on every request (nothing cached/stored).
"""
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from django.utils import timezone

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5233", end_color="2F5233", fill_type="solid")


def _new_sheet(wb, title, headers, first=False):
    ws = wb.active if first else wb.create_sheet()
    ws.title = title[:31]  # Excel's hard sheet-name length limit
    ws.sheet_view.rightToLeft = True
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    return ws


def _autosize(ws, min_width=12, max_width=45):
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(length + 2, min_width), max_width)


def _fmt_dt(dt):
    if not dt:
        return ""
    local = timezone.localtime(dt) if timezone.is_aware(dt) else dt
    return local.strftime("%Y-%m-%d %H:%M")


def _fmt_date(d):
    return d.strftime("%Y-%m-%d") if d else ""


def _who(user):
    return user.full_name if user else ""


def _bool(v):
    return "نعم" if v else "لا"


def _list_join(items):
    return "، ".join(items) if items else ""


def _dict_join(d):
    return "، ".join(f"{k}: {v}" for k, v in d.items()) if d else ""


def build_patient_workbook(patient):
    wb = openpyxl.Workbook()

    # ---- 1. المعلومات الشخصية ----
    ws = _new_sheet(wb, "المعلومات الشخصية", ["الحقل", "القيمة"], first=True)
    for row in [
        ("رقم الملف", patient.file_number),
        ("الاسم الكامل", patient.user.full_name),
        ("العمر", patient.age),
        ("الجنس", patient.gender),
        ("رقم الهاتف", patient.user.phone or ""),
        ("البريد الإلكتروني", patient.user.email),
        ("العنوان", patient.address),
        ("المهنة", patient.occupation),
        ("تاريخ الزيارة الأولى (تسجيل الملف)", _fmt_dt(patient.created_at)),
        ("تاريخ أول فتح للطبيب", _fmt_dt(patient.doctor_first_opened_at)),
        ("اللغة المفضلة", patient.get_preferred_language_display()),
        ("موعد المتابعة القادم", _fmt_dt(patient.next_followup_date)),
    ]:
        ws.append(row)
    _autosize(ws)

    # ---- 2. القياسات والهدف العلاجي ----
    a = getattr(patient, "assessment", None)
    ws = _new_sheet(wb, "القياسات والهدف العلاجي", ["الحقل", "القيمة"])
    if a:
        for row in [
            ("تاريخ الزيارة", _fmt_dt(a.visit_date)),
            ("وصل المريض؟", _bool(a.checked_in)),
            ("تم حجز الموعد؟", _bool(a.appointment_booked)),
            ("الوزن (كغم)", a.weight),
            ("الطول (م)", a.height),
            ("BMI", a.bmi),
            ("تصنيف BMI", a.bmi_class),
            ("محيط الخصر", a.waist),
            ("محيط الورك", a.hip),
            ("WHR", a.whr),
            ("تصنيف WHR", a.whr_class),
            ("الهدف", a.goal_type),
            ("الوزن الحالي (ثابت)", a.current_weight),
            ("الوزن المستهدف", a.target_weight),
            ("المدة المتوقعة", a.goal_duration),
            ("الاستمارة محفوظة نهائياً؟", _bool(a.is_submitted)),
            ("آخر تحديث", _fmt_dt(a.updated_at)),
        ]:
            ws.append(row)
    _autosize(ws)

    # ---- 3. التاريخ الطبي ونمط الحياة ----
    ws = _new_sheet(wb, "التاريخ الطبي ونمط الحياة", ["الحقل", "القيمة"])
    if a:
        for row in [
            ("الأمراض", _list_join(a.medical_history)),
            ("أمراض أخرى", a.medical_other),
            ("عمليات جراحية", a.surgeries),
            ("حساسية غذائية", a.food_allergy),
            ("مشاكل هضمية", _list_join(a.digestive_issues)),
            ("الأدوية الحالية", a.current_medications),
            ("أدوية إنقاص الوزن", _list_join(a.weight_loss_meds)),
            ("أخرى (أدوية إنقاص وزن)", a.weight_loss_meds_other),
            ("المكملات", a.supplements),
            ("النشاط البدني", a.activity_level),
            ("نوع الرياضة", a.sport_type),
            ("أيام الرياضة/أسبوع", a.sport_days_per_week),
            ("ساعات النوم", a.sleep_hours),
            ("جودة النوم", a.sleep_quality),
            ("التوتر", a.stress_level),
            ("الشهية", a.appetite),
            ("الجوع الليلي", _bool(a.night_hunger)),
            ("اشتهاء السكريات", _bool(a.sugar_craving)),
            ("مقاومة الإنسولين (تقييم سريع)", _bool(a.insulin_resistance)),
            ("أعراض هرمونية", _bool(a.hormonal_symptoms)),
            ("عدد الوجبات/يوم", a.meals_per_day),
            ("سناك", _bool(a.snack)),
            ("نمط الأكل", a.eating_type),
            ("أطعمة مفضلة", a.favorite_foods),
            ("أطعمة غير مفضلة", a.disliked_foods),
            ("الماء (لتر)", a.water_liters),
            ("القهوة/يوم", a.coffee_per_day),
            ("استهلاك السكريات", a.sugar_intake),
        ]:
            ws.append(row)
    _autosize(ws)

    # ---- 4. ملف المتابعة ----
    followup = getattr(patient, "followup", None)
    ws = _new_sheet(wb, "ملف المتابعة", ["الحقل", "القيمة"])
    if followup:
        for row in [
            ("نتائج التحاليل", _dict_join(followup.lab_results)),
            ("نوع النظام الغذائي", followup.diet_type),
            ("تفاصيل النظام الغذائي", followup.diet_details),
            ("سعرات النظام الغذائي", followup.diet_calories),
            ("الإبر", _list_join(followup.treatment_injections)),
            ("أدوية ومكملات", followup.treatment_medications),
            ("جلسات تكسير الشحم", _bool(followup.treatment_fat_burning_sessions)),
            ("مدة المتابعة القادمة", f"{followup.followup_interval_value or ''} {followup.followup_interval_unit}".strip()),
            ("غرض المتابعة", _list_join(followup.followup_purpose)),
            ("أُنشئ/عُدّل بواسطة", _who(followup.created_by)),
            ("آخر تحديث", _fmt_dt(followup.updated_at)),
        ]:
            ws.append(row)
    _autosize(ws)

    # ---- 5. متابعة التقدم ----
    ws = _new_sheet(wb, "متابعة التقدم", ["التاريخ", "الوزن", "BMI", "الالتزام", "ملاحظات", "أُدخل بواسطة"])
    for e in patient.progress_entries.select_related("created_by").order_by("-date"):
        ws.append([_fmt_dt(e.date), e.weight, e.bmi, e.commitment, e.notes, _who(e.created_by)])
    _autosize(ws)

    # ---- 6. جرعات مونجارو ----
    ws = _new_sheet(wb, "جرعات مونجارو", ["التاريخ", "الوزن", "الجرعة (ملغم)", "ملاحظات", "أُدخل بواسطة"])
    for e in patient.mounjaro_doses.select_related("created_by").all():
        ws.append([_fmt_dt(e.date), e.weight, e.dose_mg, e.notes, _who(e.created_by)])
    _autosize(ws)

    # ---- 7. جرعات أوزمبك ----
    ws = _new_sheet(wb, "جرعات أوزمبك", ["التاريخ", "الوزن", "الجرعة (ملغم)", "تركيز القلم", "ملاحظات", "أُدخل بواسطة"])
    for e in patient.ozempic_doses.select_related("created_by").all():
        ws.append([_fmt_dt(e.date), e.weight, e.dose_mg, e.get_pen_strength_display() if e.pen_strength else "", e.notes, _who(e.created_by)])
    _autosize(ws)

    # ---- 8. متابعة التحاليل ----
    ws = _new_sheet(wb, "متابعة التحاليل", ["التاريخ", "نتائج التحاليل", "ملاحظات أخرى", "أُدخل بواسطة"])
    for e in patient.lab_test_entries.select_related("created_by").order_by("-date"):
        ws.append([_fmt_dt(e.date), _dict_join(e.lab_results), e.other_notes, _who(e.created_by)])
    _autosize(ws)

    # ---- 9. الوصفات الطبية ----
    ws = _new_sheet(wb, "الوصفات الطبية", [
        "تاريخ الوصفة", "الدواء/المكمل", "الجرعة", "طريقة الاستخدام", "التكرار", "التوقيت",
        "المدة", "تاريخ البدء", "تاريخ الانتهاء", "الكمية", "الحالة العلاجية", "التعليمات",
        "ملاحظات الصنف", "ملاحظات عامة للوصفة", "كُتبت بواسطة",
    ])
    prescriptions = patient.prescriptions.select_related("created_by").prefetch_related(
        "items", "items__medication", "items__medication_dose"
    )
    for pres in prescriptions:
        items = list(pres.items.all())
        if not items:
            ws.append([_fmt_dt(pres.prescription_date), "", "", "", "", "", "", "", "", "", "", "", "", pres.general_notes, _who(pres.created_by)])
            continue
        for it in items:
            dose = it.medication_dose.display_name if it.medication_dose_id else it.custom_dose
            duration = f"{it.duration_value or ''} {it.duration_unit}".strip()
            ws.append([
                _fmt_dt(pres.prescription_date), it.display_name(), dose, it.route, it.frequency, it.timing,
                duration, _fmt_date(it.start_date), _fmt_date(it.end_date), it.quantity, it.treatment_status,
                it.instructions, it.notes, pres.general_notes, _who(pres.created_by),
            ])
    _autosize(ws)

    # ---- 10. الخطط الغذائية ----
    ws = _new_sheet(wb, "الخطط الغذائية", [
        "الاسم", "الحالة", "الإصدار", "تاريخ البدء", "المدة", "هدف العلاج",
        "مستوى النشاط", "BMR", "TDEE", "السعرات المستهدفة", "سبب الاستهداف",
        "بروتين %", "كارب %", "دهون %", "ملاحظات للطبيب", "ملاحظات للمريض",
        "أُنشئت بواسطة", "تاريخ الاعتماد", "تاريخ الإنشاء",
    ])
    plans = list(patient.nutrition_plans.select_related("created_by").prefetch_related("meals", "meals__items", "meals__items__food"))
    for plan in plans:
        duration = f"{plan.duration_value or ''} {plan.duration_unit}".strip()
        ws.append([
            plan.name, plan.get_status_display(), plan.version, _fmt_date(plan.start_date), duration,
            plan.treatment_objective, plan.activity_level, plan.bmr, plan.tdee, plan.calorie_target,
            plan.target_reason, plan.protein_pct, plan.carbs_pct, plan.fat_pct, plan.plan_notes,
            plan.patient_notes, _who(plan.created_by), _fmt_dt(plan.approved_at), _fmt_dt(plan.created_at),
        ])
    _autosize(ws)

    # ---- 11. تفاصيل الوجبات ----
    ws = _new_sheet(wb, "تفاصيل الوجبات", [
        "الخطة (اسم/إصدار)", "الوجبة", "وقت الوجبة", "الصنف", "الكمية", "الوحدة", "الحالة",
        "سعرات", "بروتين", "كارب", "دهون", "بديل", "تعليمات",
    ])
    for plan in plans:
        plan_label = f"{plan.name or 'خطة'} v{plan.version}"
        for meal in plan.meals.all():
            for item in meal.items.all():
                ws.append([
                    plan_label, meal.get_meal_type_display(), meal.time.strftime("%H:%M") if meal.time else "",
                    item.display_name(), item.quantity, item.unit, item.food_state,
                    item.calories, item.protein, item.carbs, item.fat, item.alternative_text, item.instructions,
                ])
    _autosize(ws)

    # ---- 12. ملاحظات الطبيب ----
    ws = _new_sheet(wb, "ملاحظات الطبيب", ["التاريخ", "الملاحظة", "أُدخلت بواسطة"])
    for n in patient.notes.select_related("created_by").all():
        ws.append([_fmt_dt(n.created_at), n.note, _who(n.created_by)])
    _autosize(ws)

    # ---- 13. ملاحظات الحالة الصحية ----
    ws = _new_sheet(wb, "ملاحظات الحالة الصحية", ["التاريخ", "الملاحظة", "أُدخلت بواسطة"])
    for n in patient.health_status_notes.select_related("created_by").all():
        ws.append([_fmt_dt(n.created_at), n.note, _who(n.created_by)])
    _autosize(ws)

    # ---- 14. الفواتير ----
    ws = _new_sheet(wb, "الفواتير", [
        "رقم الفاتورة", "التاريخ", "نسبة الخصم %", "سبب الخصم", "طريقة الدفع",
        "المبلغ المدفوع", "حالة الدفع", "ملاحظات", "أُنشئت بواسطة",
    ])
    invoices = list(patient.invoices.select_related("created_by").prefetch_related("items"))
    for inv in invoices:
        discount_reason = inv.discount_reason_custom or (inv.get_discount_reason_key_display() if inv.discount_reason_key else "")
        ws.append([
            inv.invoice_number, _fmt_dt(inv.created_at), inv.discount_pct, discount_reason,
            inv.payment_method, inv.amount_paid, inv.payment_status, inv.notes, _who(inv.created_by),
        ])
    _autosize(ws)

    # ---- 15. تفاصيل الفواتير ----
    ws = _new_sheet(wb, "تفاصيل الفواتير", ["رقم الفاتورة", "التاريخ", "الصنف", "سعر الوحدة", "الكمية", "الإجمالي", "متابعة مجانية؟"])
    for inv in invoices:
        for it in inv.items.all():
            ws.append([inv.invoice_number, _fmt_dt(inv.created_at), it.item_name, it.unit_price, it.quantity, it.line_total(), _bool(it.is_free_followup)])
    _autosize(ws)

    return wb


def build_all_patients_workbook(patients):
    """Same idea as build_patient_workbook, but for many patients at once —
    every sheet becomes "one row per record across ALL selected patients"
    instead of "one row per field for one patient", with رقم الملف/الاسم
    columns added so rows stay traceable to their patient. Used by the
    /admin "تصدير كل البيانات المحددة" action on PatientAdmin (see
    clinic/admin.py) — /admin already requires Django staff/superuser login,
    so no extra doctor-role check is needed here the way the single-patient
    export needs IsDoctor.

    Login credentials: only رقم الهاتف (phone, the actual login identifier)
    is included — passwords are one-way hashed and can never be exported in
    a usable form, by design (see User.set_password/AbstractUser).
    """
    patients = list(
        patients.select_related("user", "assessment", "followup")
        .prefetch_related(
            "progress_entries__created_by",
            "mounjaro_doses__created_by",
            "ozempic_doses__created_by",
            "lab_test_entries__created_by",
            "prescriptions__created_by", "prescriptions__items__medication", "prescriptions__items__medication_dose",
            "nutrition_plans__created_by", "nutrition_plans__meals__items__food",
            "notes__created_by",
            "health_status_notes__created_by",
            "invoices__created_by", "invoices__items",
        )
    )

    def pid(p):
        return [p.file_number, p.user.full_name]

    wb = openpyxl.Workbook()

    # ---- 1. المرضى ----
    ws = _new_sheet(wb, "المرضى", [
        "رقم الملف", "الاسم الكامل", "رقم الهاتف (تسجيل الدخول)", "العمر", "الجنس",
        "العنوان", "المهنة", "اللغة المفضلة", "تاريخ التسجيل", "تاريخ أول فتح للطبيب",
        "موعد المتابعة القادم",
    ], first=True)
    for p in patients:
        ws.append([
            p.file_number, p.user.full_name, p.user.phone or "", p.age, p.gender,
            p.address, p.occupation, p.get_preferred_language_display(),
            _fmt_dt(p.created_at), _fmt_dt(p.doctor_first_opened_at), _fmt_dt(p.next_followup_date),
        ])
    _autosize(ws)

    # ---- 2. القياسات والهدف العلاجي ----
    ws = _new_sheet(wb, "القياسات والهدف العلاجي", [
        "رقم الملف", "الاسم", "تاريخ الزيارة", "وصل المريض؟", "تم حجز الموعد؟",
        "الوزن (كغم)", "الطول (م)", "BMI", "تصنيف BMI", "محيط الخصر", "محيط الورك",
        "WHR", "تصنيف WHR", "الهدف", "الوزن الحالي (ثابت)", "الوزن المستهدف",
        "المدة المتوقعة", "الاستمارة محفوظة نهائياً؟", "آخر تحديث",
    ])
    for p in patients:
        a = getattr(p, "assessment", None)
        if not a:
            continue
        ws.append(pid(p) + [
            _fmt_dt(a.visit_date), _bool(a.checked_in), _bool(a.appointment_booked),
            a.weight, a.height, a.bmi, a.bmi_class, a.waist, a.hip, a.whr, a.whr_class,
            a.goal_type, a.current_weight, a.target_weight, a.goal_duration,
            _bool(a.is_submitted), _fmt_dt(a.updated_at),
        ])
    _autosize(ws)

    # ---- 3. التاريخ الطبي ونمط الحياة ----
    ws = _new_sheet(wb, "التاريخ الطبي ونمط الحياة", [
        "رقم الملف", "الاسم", "الأمراض", "أمراض أخرى", "عمليات جراحية", "حساسية غذائية",
        "مشاكل هضمية", "الأدوية الحالية", "أدوية إنقاص الوزن", "أخرى (أدوية إنقاص وزن)",
        "المكملات", "النشاط البدني", "نوع الرياضة", "أيام الرياضة/أسبوع", "ساعات النوم",
        "جودة النوم", "التوتر", "الشهية", "الجوع الليلي", "اشتهاء السكريات",
        "مقاومة الإنسولين (تقييم سريع)", "أعراض هرمونية", "عدد الوجبات/يوم", "سناك",
        "نمط الأكل", "أطعمة مفضلة", "أطعمة غير مفضلة", "الماء (لتر)", "القهوة/يوم", "استهلاك السكريات",
    ])
    for p in patients:
        a = getattr(p, "assessment", None)
        if not a:
            continue
        ws.append(pid(p) + [
            _list_join(a.medical_history), a.medical_other, a.surgeries, a.food_allergy,
            _list_join(a.digestive_issues), a.current_medications, _list_join(a.weight_loss_meds),
            a.weight_loss_meds_other, a.supplements, a.activity_level, a.sport_type,
            a.sport_days_per_week, a.sleep_hours, a.sleep_quality, a.stress_level, a.appetite,
            _bool(a.night_hunger), _bool(a.sugar_craving), _bool(a.insulin_resistance),
            _bool(a.hormonal_symptoms), a.meals_per_day, _bool(a.snack), a.eating_type,
            a.favorite_foods, a.disliked_foods, a.water_liters, a.coffee_per_day, a.sugar_intake,
        ])
    _autosize(ws)

    # ---- 4. ملف المتابعة ----
    ws = _new_sheet(wb, "ملف المتابعة", [
        "رقم الملف", "الاسم", "نتائج التحاليل", "نوع النظام الغذائي", "تفاصيل النظام الغذائي",
        "سعرات النظام الغذائي", "الإبر", "أدوية ومكملات", "جلسات تكسير الشحم",
        "مدة المتابعة القادمة", "غرض المتابعة", "أُنشئ/عُدّل بواسطة", "آخر تحديث",
    ])
    for p in patients:
        followup = getattr(p, "followup", None)
        if not followup:
            continue
        ws.append(pid(p) + [
            _dict_join(followup.lab_results), followup.diet_type, followup.diet_details,
            followup.diet_calories, _list_join(followup.treatment_injections),
            followup.treatment_medications, _bool(followup.treatment_fat_burning_sessions),
            f"{followup.followup_interval_value or ''} {followup.followup_interval_unit}".strip(),
            _list_join(followup.followup_purpose), _who(followup.created_by), _fmt_dt(followup.updated_at),
        ])
    _autosize(ws)

    # ---- 5. متابعة التقدم ----
    ws = _new_sheet(wb, "متابعة التقدم", ["رقم الملف", "الاسم", "التاريخ", "الوزن", "BMI", "الالتزام", "ملاحظات", "أُدخل بواسطة"])
    for p in patients:
        for e in p.progress_entries.all():
            ws.append(pid(p) + [_fmt_dt(e.date), e.weight, e.bmi, e.commitment, e.notes, _who(e.created_by)])
    _autosize(ws)

    # ---- 6. جرعات مونجارو ----
    ws = _new_sheet(wb, "جرعات مونجارو", ["رقم الملف", "الاسم", "التاريخ", "الوزن", "الجرعة (ملغم)", "ملاحظات", "أُدخل بواسطة"])
    for p in patients:
        for e in p.mounjaro_doses.all():
            ws.append(pid(p) + [_fmt_dt(e.date), e.weight, e.dose_mg, e.notes, _who(e.created_by)])
    _autosize(ws)

    # ---- 7. جرعات أوزمبك ----
    ws = _new_sheet(wb, "جرعات أوزمبك", ["رقم الملف", "الاسم", "التاريخ", "الوزن", "الجرعة (ملغم)", "تركيز القلم", "ملاحظات", "أُدخل بواسطة"])
    for p in patients:
        for e in p.ozempic_doses.all():
            ws.append(pid(p) + [
                _fmt_dt(e.date), e.weight, e.dose_mg,
                e.get_pen_strength_display() if e.pen_strength else "", e.notes, _who(e.created_by),
            ])
    _autosize(ws)

    # ---- 8. متابعة التحاليل ----
    ws = _new_sheet(wb, "متابعة التحاليل", ["رقم الملف", "الاسم", "التاريخ", "نتائج التحاليل", "ملاحظات أخرى", "أُدخل بواسطة"])
    for p in patients:
        for e in p.lab_test_entries.all():
            ws.append(pid(p) + [_fmt_dt(e.date), _dict_join(e.lab_results), e.other_notes, _who(e.created_by)])
    _autosize(ws)

    # ---- 9. الوصفات الطبية ----
    ws = _new_sheet(wb, "الوصفات الطبية", [
        "رقم الملف", "الاسم", "تاريخ الوصفة", "الدواء/المكمل", "الجرعة", "طريقة الاستخدام",
        "التكرار", "التوقيت", "المدة", "تاريخ البدء", "تاريخ الانتهاء", "الكمية",
        "الحالة العلاجية", "التعليمات", "ملاحظات الصنف", "ملاحظات عامة للوصفة", "كُتبت بواسطة",
    ])
    for p in patients:
        for pres in p.prescriptions.all():
            items = list(pres.items.all())
            if not items:
                ws.append(pid(p) + [_fmt_dt(pres.prescription_date), "", "", "", "", "", "", "", "", "", "", "", "", pres.general_notes, _who(pres.created_by)])
                continue
            for it in items:
                dose = it.medication_dose.display_name if it.medication_dose_id else it.custom_dose
                duration = f"{it.duration_value or ''} {it.duration_unit}".strip()
                ws.append(pid(p) + [
                    _fmt_dt(pres.prescription_date), it.display_name(), dose, it.route, it.frequency, it.timing,
                    duration, _fmt_date(it.start_date), _fmt_date(it.end_date), it.quantity, it.treatment_status,
                    it.instructions, it.notes, pres.general_notes, _who(pres.created_by),
                ])
    _autosize(ws)

    # ---- 10. الخطط الغذائية ----
    ws = _new_sheet(wb, "الخطط الغذائية", [
        "رقم الملف", "الاسم", "اسم الخطة", "الحالة", "الإصدار", "تاريخ البدء", "المدة",
        "هدف العلاج", "مستوى النشاط", "BMR", "TDEE", "السعرات المستهدفة", "سبب الاستهداف",
        "بروتين %", "كارب %", "دهون %", "ملاحظات للطبيب", "ملاحظات للمريض",
        "أُنشئت بواسطة", "تاريخ الاعتماد", "تاريخ الإنشاء",
    ])
    for p in patients:
        for plan in p.nutrition_plans.all():
            duration = f"{plan.duration_value or ''} {plan.duration_unit}".strip()
            ws.append(pid(p) + [
                plan.name, plan.get_status_display(), plan.version, _fmt_date(plan.start_date), duration,
                plan.treatment_objective, plan.activity_level, plan.bmr, plan.tdee, plan.calorie_target,
                plan.target_reason, plan.protein_pct, plan.carbs_pct, plan.fat_pct, plan.plan_notes,
                plan.patient_notes, _who(plan.created_by), _fmt_dt(plan.approved_at), _fmt_dt(plan.created_at),
            ])
    _autosize(ws)

    # ---- 11. تفاصيل الوجبات ----
    ws = _new_sheet(wb, "تفاصيل الوجبات", [
        "رقم الملف", "الاسم", "الخطة (اسم/إصدار)", "الوجبة", "وقت الوجبة", "الصنف",
        "الكمية", "الوحدة", "الحالة", "سعرات", "بروتين", "كارب", "دهون", "بديل", "تعليمات",
    ])
    for p in patients:
        for plan in p.nutrition_plans.all():
            plan_label = f"{plan.name or 'خطة'} v{plan.version}"
            for meal in plan.meals.all():
                for item in meal.items.all():
                    ws.append(pid(p) + [
                        plan_label, meal.get_meal_type_display(), meal.time.strftime("%H:%M") if meal.time else "",
                        item.display_name(), item.quantity, item.unit, item.food_state,
                        item.calories, item.protein, item.carbs, item.fat, item.alternative_text, item.instructions,
                    ])
    _autosize(ws)

    # ---- 12. ملاحظات الطبيب ----
    ws = _new_sheet(wb, "ملاحظات الطبيب", ["رقم الملف", "الاسم", "التاريخ", "الملاحظة", "أُدخلت بواسطة"])
    for p in patients:
        for n in p.notes.all():
            ws.append(pid(p) + [_fmt_dt(n.created_at), n.note, _who(n.created_by)])
    _autosize(ws)

    # ---- 13. ملاحظات الحالة الصحية ----
    ws = _new_sheet(wb, "ملاحظات الحالة الصحية", ["رقم الملف", "الاسم", "التاريخ", "الملاحظة", "أُدخلت بواسطة"])
    for p in patients:
        for n in p.health_status_notes.all():
            ws.append(pid(p) + [_fmt_dt(n.created_at), n.note, _who(n.created_by)])
    _autosize(ws)

    # ---- 14. الفواتير ----
    ws = _new_sheet(wb, "الفواتير", [
        "رقم الملف", "الاسم", "رقم الفاتورة", "التاريخ", "نسبة الخصم %", "سبب الخصم",
        "طريقة الدفع", "المبلغ المدفوع", "حالة الدفع", "ملاحظات", "أُنشئت بواسطة",
    ])
    for p in patients:
        for inv in p.invoices.all():
            discount_reason = inv.discount_reason_custom or (inv.get_discount_reason_key_display() if inv.discount_reason_key else "")
            ws.append(pid(p) + [
                inv.invoice_number, _fmt_dt(inv.created_at), inv.discount_pct, discount_reason,
                inv.payment_method, inv.amount_paid, inv.payment_status, inv.notes, _who(inv.created_by),
            ])
    _autosize(ws)

    # ---- 15. تفاصيل الفواتير ----
    ws = _new_sheet(wb, "تفاصيل الفواتير", [
        "رقم الملف", "الاسم", "رقم الفاتورة", "التاريخ", "الصنف", "سعر الوحدة",
        "الكمية", "الإجمالي", "متابعة مجانية؟",
    ])
    for p in patients:
        for inv in p.invoices.all():
            for it in inv.items.all():
                ws.append(pid(p) + [
                    inv.invoice_number, _fmt_dt(inv.created_at), it.item_name, it.unit_price,
                    it.quantity, it.line_total(), _bool(it.is_free_followup),
                ])
    _autosize(ws)

    return wb
